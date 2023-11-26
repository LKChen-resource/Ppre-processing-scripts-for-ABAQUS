from abaqus import *
from abaqusConstants import *
from odbAccess import *
from textRepr  import *
import regionToolset
import openpyxl as op
filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\CFST_Parameter_Analyse.xlsx'
wb=op.load_workbook(filename)
sh=wb["Sheet7"]
for i in range(2,27): #excel row-1  : end row **********input*******
    #-----Create a new model
    ModleNameTemp1=str(sh.cell(row=i,column=0).value)
    ModleName=ModleNameTemp1.replace('.','_')
    OuterDimension=sh.cell(row=i,column=3).value
    Tickness=sh.cell(row=i,column=4).value
    TLen=sh.cell(row=i,column=6).value  #TubeLength
    OR=OuterDimension/2.  #OuterRadius
    IR=OR-Tickness #interRadius



    Fys=sh.cell(row=i,column=2).value  #Steel yield strength
    Fc=sh.cell(row=i,column=1).value  #Concrete strength
    ft=round(0.3*(Fc-8.0)**(2/3.),2)


    Es=200000 #steel elastic Modulus
    Econ=round(4730*Fc**0.5,0)  #Concrete elastic Modulus

    mdb.Model(name=ModleName, modelType=STANDARD_EXPLICIT)
    #: The model "C1" has been created.
    #session.viewports['Viewport: 1'].setValues(displayedObject=None)
    #-----Create  Steel  Part
    s = mdb.models[ModleName].ConstrainedSketch(name='__profile__', sheetSize=200.0)
    g, v, d, c = s.geometry, s.vertices, s.dimensions, s.constraints
    s.setPrimaryObject(option=STANDALONE)

    s.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(0.0, OR))#OuterRadius
    s.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(0.0, IR))#interRadius
    p = mdb.models[ModleName].Part(name='Steel', dimensionality=THREE_D, 
        type=DEFORMABLE_BODY)
    p = mdb.models[ModleName].parts['Steel']
    p.BaseSolidExtrude(sketch=s, depth=TLen) #Model  Length
    s.unsetPrimaryObject()
    p = mdb.models[ModleName].parts['Steel']
    #session.viewports['Viewport: 1'].setValues(displayedObject=p)
    del mdb.models[ModleName].sketches['__profile__']

    #----Create Concrete Part
    s1 = mdb.models[ModleName].ConstrainedSketch(name='__profile__', sheetSize=200.0)
    g, v, d, c = s1.geometry, s1.vertices, s1.dimensions, s1.constraints
    s1.setPrimaryObject(option=STANDALONE)

    s1.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(0.0, IR))
    p = mdb.models[ModleName].Part(name='Concrete', dimensionality=THREE_D, 
        type=DEFORMABLE_BODY)
    p = mdb.models[ModleName].parts['Concrete']
    p.BaseSolidExtrude(sketch=s1, depth=TLen) # Model  Length
    s1.unsetPrimaryObject()
    p = mdb.models[ModleName].parts['Concrete']
    #session.viewports['Viewport: 1'].setValues(displayedObject=p)
    del mdb.models[ModleName].sketches['__profile__']

    #create  rigid plate
    s1 = mdb.models[ModleName].ConstrainedSketch(name='__profile__', 
        sheetSize=200.0)
    g, v, d, c = s1.geometry, s1.vertices, s1.dimensions, s1.constraints
    s1.setPrimaryObject(option=STANDALONE)
    s1.rectangle(point1=(-200.0, -200.0), point2=(200.0, 200.0))
    p = mdb.models[ModleName].Part(name='plate', dimensionality=THREE_D, 
        type=DISCRETE_RIGID_SURFACE)
    p = mdb.models[ModleName].parts['plate']
    p.BaseSolidExtrude(sketch=s1, depth=30.0)
    s1.unsetPrimaryObject()
    ##p = mdb.models[ModleName].parts['plate']
    del mdb.models[ModleName].sketches['__profile__']

    #transfor to shell
    ##p = mdb.models[ModleName].parts['plate']
    c1 = p.cells
    p.RemoveCells(cellList = c1[0:1])
    #: 
    #: One shell per selected cell has been created from the cell's faces.


    #----Property-----------------------------------------------------------------
    mdb.models[ModleName].Material(name='Steel')
    mdb.models[ModleName].materials['Steel'].Density(table=((7.85e-09, ), )) #
    mdb.models[ModleName].materials['Steel'].Elastic(table=((Es, 0.3), )) #
    mdb.models[ModleName].materials['Steel'].Plastic(table=((Fys, 0.0),(Fys+100,0.1)))  #

    mdb.models[ModleName].HomogeneousSolidSection(name='Section-1', material='Steel',
    thickness=None)
        #defined Property
    p = mdb.models[ModleName].parts['Steel']
    c = p.cells
    cells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    region = regionToolset.Region(cells=cells)
    p = mdb.models[ModleName].parts['Steel']
    p.SectionAssignment(region=region, sectionName='Section-1', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)

    #Concrete
    mdb.models[ModleName].Material(name='Concrete')
    mdb.models[ModleName].materials['Concrete'].Density(table=((2.4e-09, ), ))
    mdb.models[ModleName].materials['Concrete'].Depvar(deleteVar=28, n=49)
    mdb.models[ModleName].materials['Concrete'].UserMaterial(mechanicalConstants=(1.0, 
        Econ, 0.2, Fc, ft, round(0.073*Fc**0.18,3), 0.0, 0.08, 0.003, 2.0, 1e-06, 0.01, 15.0))

    mdb.models[ModleName].HomogeneousSolidSection(name='Section-2', material='Concrete', 
        thickness=None)

    p = mdb.models[ModleName].parts['Concrete']
    c = p.cells
    cells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    region = regionToolset.Region(cells=cells)
    p = mdb.models[ModleName].parts['Concrete']
    p.SectionAssignment(region=region, sectionName='Section-2', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)


    #Assmebly
    a = mdb.models[ModleName].rootAssembly
    a.DatumCsysByDefault(CARTESIAN)
    p = mdb.models[ModleName].parts['Concrete']
    a.Instance(name='Concrete-1', part=p, dependent=ON)
    p = mdb.models[ModleName].parts['Steel']
    a.Instance(name='Steel-1', part=p, dependent=ON)
 
    #Step
    #Creat dynamic step
    mdb.models[ModleName].ExplicitDynamicsStep(name='Step-1', previous='Initial', 
        massScaling=((SEMI_AUTOMATIC, MODEL, AT_BEGINNING, 0.0, 0.00001, BELOW_MIN, 
        0, 0, 0.0, 0.0, 0, None), ), improvedDtMethod=ON)


    #modify output
    mdb.models[ModleName].fieldOutputRequests['F-Output-1'].setValues(variables=('S', 
        'SVAVG', 'PE', 'PEVAVG', 'PEEQ', 'PEEQVAVG', 'LE', 'U', 'V', 'A', 'RF', 
        'CSTRESS', 'EVOL', 'EVF', 'SDV', 'STATUS'))
    mdb.models[ModleName].fieldOutputRequests['F-Output-1'].setValues(numIntervals=100)

    #Create RP
    a.ReferencePoint(point=(0.0, 0.0, TLen+30)) #len+30

    a.ReferencePoint(point=(0.0, 0.0, TLen+20))#len+20

    a.ReferencePoint(point=(0.0, 0.0, TLen+10))##len+10

    a.ReferencePoint(point=(0.0, 0.0, -10.0))

    #Create  hrad contact
    mdb.models[ModleName].ContactProperty('IntProp-1')
    mdb.models[ModleName].interactionProperties['IntProp-1'].TangentialBehavior(
        formulation=PENALTY, directionality=ISOTROPIC, slipRateDependency=OFF, 
        pressureDependency=OFF, temperatureDependency=OFF, dependencies=0, table=((
        0.6, ), ), shearStressLimit=None, maximumElasticSlip=FRACTION, 
        fraction=0.005, elasticSlipStiffness=None) #friction coefficient 0.05
    mdb.models[ModleName].interactionProperties['IntProp-1'].NormalBehavior(
        pressureOverclosure=HARD, allowSeparation=ON, 
        constraintEnforcementMethod=DEFAULT)
    #: The interaction property "IntProp-1" has been created.
    #creat a constraint under bottom
    a = mdb.models[ModleName].rootAssembly
    r1 = a.referencePoints
    refPoints1=(r1[9], )
    region1=regionToolset.Region(referencePoints=refPoints1)
    s1 = a.instances['Concrete-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#4 ]', ), )
    s2 = a.instances['Steel-1'].faces
    side1Faces2 = s2.getSequenceFromMask(mask=('[#8 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1+side1Faces2)
    mdb.models[ModleName].Coupling(name='Constraint-1', 
        controlPoint=region1, surface=region2, influenceRadius=WHOLE_SURFACE, 
        couplingType=KINEMATIC, localCsys=None, u1=ON, u2=ON, u3=ON, ur1=ON, 
        ur2=ON, ur3=ON)

    # RF2 control steel
    r1 = a.referencePoints
    refPoints1=(r1[7], )
    region1=regionToolset.Region(referencePoints=refPoints1)
    s1 = a.instances['Steel-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#4 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].Coupling(name='Constraint-3', controlPoint=region1, 
        surface=region2, influenceRadius=WHOLE_SURFACE, couplingType=KINEMATIC, 
        localCsys=None, u1=ON, u2=ON, u3=ON, ur1=ON, ur2=ON, ur3=ON)

    #RF3 control Conrete

    r1 = a.referencePoints
    refPoints1=(r1[8], )
    region1=regionToolset.Region(referencePoints=refPoints1)

    s1 = a.instances['Concrete-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#2 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].Coupling(name='Constraint-4', controlPoint=region1, 
        surface=region2, influenceRadius=WHOLE_SURFACE, couplingType=KINEMATIC, 
        localCsys=None, u1=ON, u2=ON, u3=ON, ur1=ON, ur2=ON, ur3=ON)

    #create surface to surface

    s1 = a.instances['Concrete-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#1 ]', ), )
    region1=regionToolset.Region(side1Faces=side1Faces1)

    s1 = a.instances['Steel-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#2 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].SurfaceToSurfaceContactExp(name ='Int-1', 
        createStepName='Initial', master = region1, slave = region2, 
        mechanicalConstraint=KINEMATIC, sliding=FINITE, 
        interactionProperty='IntProp-1', initialClearance=OMIT, datumAxis=None, 
        clearanceRegion=None)
    #: The interaction "Int-1" has been created.
    #: The interaction "Int-2" has been created.
    #: The interaction "Int-3" has been created.

    #Create Boundary
    #Bottom

    r1 = a.referencePoints
    refPoints1=(r1[9], )
    region = regionToolset.Region(referencePoints=refPoints1)
    mdb.models[ModleName].DisplacementBC(name='BC-1', createStepName='Initial', 
        region=region, u1=SET, u2=SET, u3=SET, ur1=SET, ur2=SET, ur3=SET, 
        amplitude=UNSET, distributionType=UNIFORM, fieldName='', localCsys=None)

    #upon
    mdb.models[ModleName].TabularAmplitude(name='Amp-1', timeSpan=STEP, 
        smooth=SOLVER_DEFAULT, data=((0.0, 0.0), (1.0, 1.0)))

    r1 = a.referencePoints
    refPoints1=(r1[7],r1[8] )
    region = regionToolset.Region(referencePoints=refPoints1)
    mdb.models[ModleName].DisplacementBC(name='BC-2', createStepName='Step-1', 
        region=region, u1=0.0, u2=0.0, u3=-30.0, ur1=0.0, ur2=0.0, ur3=0.0, 
        amplitude='Amp-1', fixed=OFF, distributionType=UNIFORM, fieldName='', 
        localCsys=None)

    ##Mesh
    #partition cell
    p = mdb.models[ModleName].parts['Steel']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    e, v1, d1 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e[0], point=v1[0], cells=pickedCells)
    # p = mdb.models[ModleName].parts['Steel']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#3 ]', ), )
    e1, v2, d2 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e1[9], cells=pickedCells, 
        point=p.InterestingPoint(edge=e1[9], rule=MIDDLE))

    p = mdb.models[ModleName].parts['Concrete']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    e, v1, d1 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e[0], cells=pickedCells, 
        point=p.InterestingPoint(edge=e[0], rule=MIDDLE))
    # p = mdb.models[ModleName].parts['Concrete']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#3 ]', ), )
    e1, v2, d2 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e1[4], cells=pickedCells, 
        point=p.InterestingPoint(edge=e1[4], rule=MIDDLE))
    # draw mesh
    p = mdb.models[ModleName].parts['Concrete']
    p.seedPart(size=25., deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models[ModleName].parts['Concrete']
    p.generateMesh()
    p = mdb.models[ModleName].parts['Steel']
    p.seedPart(size=25., deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models[ModleName].parts['Steel']
    e = p.edges
    pickedEdges = e.getSequenceFromMask(mask=('[#4001011 ]', ), )
    p.seedEdgeBySize(edges=pickedEdges, size=25./2, deviationFactor=0.1, 
    minSizeFactor=0.1, constraint=FINER)
    p.generateMesh()

    p = mdb.models[ModleName].parts['plate']
    p = mdb.models[ModleName].parts['plate']
    p.seedPart(size=200.0, deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models[ModleName].parts['plate']

    p.generateMesh()

    # general job
    JobName=ModleName.replace('-','_')
    mdb.Job(name=JobName, model=ModleName, description='', type=ANALYSIS, 
        atTime=None, waitMinutes=0, waitHours=0, queue=None, memory=90, 
        memoryUnits=PERCENTAGE, explicitPrecision=DOUBLE_PLUS_PACK, 
        nodalOutputPrecision=SINGLE, echoPrint=OFF, modelPrint=OFF, 
        contactPrint=OFF, historyPrint=OFF, 
        userSubroutine='E:\\DOCUMENTS\\1YueLu\\1.Subjects\\Poisson\\ABAQUSModel\\TriaxialCompression\\CDPM2\\master\\cdpm2vumat.f', 
        scratch='', resultsFormat=ODB, parallelizationMethodExplicit=DOMAIN, 
        numDomains=1, activateLoadBalancing=False, multiprocessingMode=DEFAULT, 
        numCpus=1)

    a = mdb.models[ModleName].rootAssembly
    a.regenerate()



#wb.save(filename)