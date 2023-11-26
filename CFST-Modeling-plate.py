from abaqus import *
from abaqusConstants import *
from odbAccess import *
from textRepr  import *
import regionToolset
import openpyxl as op
filename=r'C:\Users\LKChen\Desktop\CFST-BaseDatas.xlsx'
wb=op.load_workbook(filename)
sh=wb["Sheet2"]
# steel plastic five flow----------
def sigma(fyl,Es):
    ee=0.8*fyl/Es
    ee1=1.5*ee
    ee2=10*ee1        
    ee3=100*ee1
    # print(ee,ee1,ee2,ee3)
    e=[ee,1.1*ee,1.2*ee,1.3*ee,1.4*ee,ee1,ee2,ee3]
    pe=[0,round(e[1]-ee,5),round(e[2]-ee,5),round(e[3]-ee,5),round(e[4]-ee,5),
    round(e[5]-ee,5),round(e[6]-ee,5),round(e[7]-ee,5)]
    #print(e)
    M=0.2*fyl/(ee1-ee)**2
    N=2*M*ee1
    P=0.8*fyl+M*ee**2-N*ee
    sigma=[]
    for i in range(len(e)):
        if e[i] <=ee:
            sigma.append(round(Es*e[i],2))
        elif ee<e[i]<=ee1:
            sigma.append(round(-1*M*e[i]**2+N*e[i]+P,2))
        elif ee1<e[i]<=ee2:
            sigma.append(round(fyl,2))
        elif ee2<e[i]<=ee3:
            sigma.append(round(fyl*(1+0.6711*(e[i]-ee2)/(ee3-ee2)),2))
        else:
            sigma.append(round(1.6711*fyl,2))
    return sigma,pe
    #-------------

for i in range(31,37): #excel row-1  : end row **********input*******
    #-----Create a new model
    ModleName=str(sh.cell(row=i,column=0).value)
    OuterDimension=sh.cell(row=i,column=1).value
    Tickness=sh.cell(row=i,column=2).value
    TLen=sh.cell(row=i,column=3).value  #TubeLength
    OR=OuterDimension/2  #OuterRadius
    IR=OR-Tickness #interRadius

    Es=sh.cell(row=i,column=7).value #steel elastic Modulus
    Econ=sh.cell(row=i,column=9).value  #Concrete elastic Modulus

    Fys=sh.cell(row=i,column=6).value  #Steel yield strength
    Fc=sh.cell(row=i,column=8).value  #Concrete strength
    ft=Fc/10

    PlasticFlowR=sigma(Fys,Es)

    mdb.Model(name=ModleName, modelType=STANDARD_EXPLICIT)
    #: The model "C1" has been created.
    #session.viewports['Viewport: 1'].setValues(displayedObject=None)
    #-----Create  Steel  Part
    s = mdb.models[ModleName].ConstrainedSketch(name='__profile__', sheetSize=200.0)
    g, v, d, c = s.geometry, s.vertices, s.dimensions, s.constraints
    s.setPrimaryObject(option=STANDALONE)

    s.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(0.0, OR))#OuterRadius
    s.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(0.0, IR))#interRadius
    p = mdb.models[ModleName].Part(name='Steel', dimensionality=THREE_D, 
        type=DEFORMABLE_BODY)
    p = mdb.models[ModleName].parts['Steel']
    p.BaseSolidExtrude(sketch=s, depth=TLen) #Model  Length
    s.unsetPrimaryObject()
    p = mdb.models[ModleName].parts['Steel']
    #session.viewports['Viewport: 1'].setValues(displayedObject=p)
    del mdb.models[ModleName].sketches['__profile__']

    #----Create Concrete Part
    s1 = mdb.models[ModleName].ConstrainedSketch(name='__profile__', sheetSize=200.0)
    g, v, d, c = s1.geometry, s1.vertices, s1.dimensions, s1.constraints
    s1.setPrimaryObject(option=STANDALONE)

    s1.CircleByCenterPerimeter(center=(0.0, 0.0), point1=(0.0, IR))
    p = mdb.models[ModleName].Part(name='Concrete', dimensionality=THREE_D, 
        type=DEFORMABLE_BODY)
    p = mdb.models[ModleName].parts['Concrete']
    p.BaseSolidExtrude(sketch=s1, depth=TLen) # Model  Length
    s1.unsetPrimaryObject()
    p = mdb.models[ModleName].parts['Concrete']
    #session.viewports['Viewport: 1'].setValues(displayedObject=p)
    del mdb.models[ModleName].sketches['__profile__']

    #create  rigid plate
    s1 = mdb.models[ModleName].ConstrainedSketch(name='__profile__', 
        sheetSize=200.0)
    g, v, d, c = s1.geometry, s1.vertices, s1.dimensions, s1.constraints
    s1.setPrimaryObject(option=STANDALONE)
    s1.rectangle(point1=(-200.0, -200.0), point2=(200.0, 200.0))
    p = mdb.models[ModleName].Part(name='plate', dimensionality=THREE_D, 
        type=DISCRETE_RIGID_SURFACE)
    p = mdb.models[ModleName].parts['plate']
    p.BaseSolidExtrude(sketch=s1, depth=30.0)
    s1.unsetPrimaryObject()
    ##p = mdb.models[ModleName].parts['plate']
    session.viewports['Viewport: 1'].setValues(displayedObject=p)
    del mdb.models[ModleName].sketches['__profile__']

    #transfor to shell
    ##p = mdb.models[ModleName].parts['plate']
    c1 = p.cells
    p.RemoveCells(cellList = c1[0:1])
    #: 
    #: One shell per selected cell has been created from the cell's faces.


    #----Property-----------------------------------------------------------------
    mdb.models[ModleName].Material(name='Steel')
    mdb.models[ModleName].materials['Steel'].Density(table=((7.85e-09, ), )) #
    mdb.models[ModleName].materials['Steel'].Elastic(table=((Es, 0.3), )) #
    mdb.models[ModleName].materials['Steel'].Plastic(table=((PlasticFlowR[0][0], PlasticFlowR[1][0]), (
     PlasticFlowR[0][1], PlasticFlowR[1][1]),(PlasticFlowR[0][2], PlasticFlowR[1][2]),(PlasticFlowR[0][3], 
     PlasticFlowR[1][3]),(PlasticFlowR[0][4], PlasticFlowR[1][4]),
     (PlasticFlowR[0][5], PlasticFlowR[1][5]),(PlasticFlowR[0][6],
      PlasticFlowR[1][6]),(PlasticFlowR[0][7], PlasticFlowR[1][7]) ))  #

    mdb.models[ModleName].HomogeneousSolidSection(name='Section-1', material='Steel',
    thickness=None)
        #defined Property
    p = mdb.models[ModleName].parts['Steel']
    c = p.cells
    cells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    region = regionToolset.Region(cells=cells)
    p = mdb.models[ModleName].parts['Steel']
    p.SectionAssignment(region=region, sectionName='Section-1', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)

    #Concrete
    mdb.models[ModleName].Material(name='Concrete')
    mdb.models[ModleName].materials['Concrete'].Density(table=((2.4e-09, ), ))
    mdb.models[ModleName].materials['Concrete'].Depvar(deleteVar=28, n=49)
    mdb.models[ModleName].materials['Concrete'].UserMaterial(mechanicalConstants=(1.0, 
        Econ, 0.2, Fc, ft, 0.01, 0.0, 0.08, 0.003, 2.0, 1e-06, 0.01, 15.0))

    mdb.models[ModleName].HomogeneousSolidSection(name='Section-2', material='Concrete', 
        thickness=None)

    p = mdb.models[ModleName].parts['Concrete']
    c = p.cells
    cells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    region = regionToolset.Region(cells=cells)
    p = mdb.models[ModleName].parts['Concrete']
    p.SectionAssignment(region=region, sectionName='Section-2', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)


    #Assmebly
    a = mdb.models[ModleName].rootAssembly
    a.DatumCsysByDefault(CARTESIAN)
    p = mdb.models[ModleName].parts['Concrete']
    a.Instance(name='Concrete-1', part=p, dependent=ON)
    p = mdb.models[ModleName].parts['Steel']
    a.Instance(name='Steel-1', part=p, dependent=ON)

    #a = mdb.models[ModleName].rootAssembly
    p = mdb.models[ModleName].parts['plate']
    a.Instance(name='plate-1', part=p, dependent=ON)
    #a = mdb.models[ModleName].rootAssembly
    a.translate(instanceList=('plate-1', ), vector=(0.0, 0.0, -30.0))
    #: The instance plate-1 was translated by 0., 0., -30. with respect to the assembly coordinate system
    a1 = mdb.models[ModleName].rootAssembly
    a1.LinearInstancePattern(instanceList=('plate-1', ), direction1=(0.0, 0.0, 
        1.0), direction2=(0.0, 1.0, 0.0), number1=2, number2=1, spacing1=TLen+30, 
        spacing2=400.0)





    #Step
    #Creat dynamic step
    mdb.models[ModleName].ExplicitDynamicsStep(name='Step-1', previous='Initial', 
        massScaling=((SEMI_AUTOMATIC, MODEL, AT_BEGINNING, 0.0, 0.0001, BELOW_MIN, 
        0, 0, 0.0, 0.0, 0, None), ), improvedDtMethod=ON)


    #modify output
    mdb.models[ModleName].fieldOutputRequests['F-Output-1'].setValues(variables=('S', 
        'SVAVG', 'PE', 'PEVAVG', 'PEEQ', 'PEEQVAVG', 'LE', 'U', 'V', 'A', 'RF', 
        'CSTRESS', 'EVOL', 'EVF', 'SDV', 'STATUS'))
    mdb.models[ModleName].fieldOutputRequests['F-Output-1'].setValues(numIntervals=100)

    #Create RP
    a.ReferencePoint(point=(0.0, 0.0, TLen+50)) #len+30

    #a.ReferencePoint(point=(0.0, 0.0, TLen+20))#len+20

    #a.ReferencePoint(point=(0.0, 0.0, TLen+10))##len+10

    a.ReferencePoint(point=(0.0, 0.0, -40.0))

    #Create  hrad contact
    mdb.models[ModleName].ContactProperty('IntProp-1')
    mdb.models[ModleName].interactionProperties['IntProp-1'].TangentialBehavior(
        formulation=PENALTY, directionality=ISOTROPIC, slipRateDependency=OFF, 
        pressureDependency=OFF, temperatureDependency=OFF, dependencies=0, table=((
        0.05, ), ), shearStressLimit=None, maximumElasticSlip=FRACTION, 
        fraction=0.005, elasticSlipStiffness=None) #friction coefficient 0.05
    mdb.models[ModleName].interactionProperties['IntProp-1'].NormalBehavior(
        pressureOverclosure=HARD, allowSeparation=ON, 
        constraintEnforcementMethod=DEFAULT)
    #: The interaction property "IntProp-1" has been created.

    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['plate-1-lin-2-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#20 ]', ), )
    region1=a.Surface(side1Faces=side1Faces1, name='m_Surf-1')
    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['Steel-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#4 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].SurfaceToSurfaceContactExp(name ='Int-1', 
        createStepName='Initial', master = region1, slave = region2, 
        mechanicalConstraint=KINEMATIC, sliding=FINITE, 
        interactionProperty='IntProp-1', initialClearance=OMIT, datumAxis=None, 
        clearanceRegion=None)
    #: The interaction "Int-1" has been created.
    mdb.models[ModleName].Interaction(name='Int-1-Copy', 
        objectToCopy=mdb.models[ModleName].interactions['Int-1'], 
        toStepName='Initial')
    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['Concrete-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#2 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].interactions['Int-1-Copy'].setValues(slave=region2, 
        mechanicalConstraint=KINEMATIC, sliding=FINITE, 
        interactionProperty='IntProp-1', initialClearance=OMIT, datumAxis=None, 
        clearanceRegion=None)

    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['plate-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#10 ]', ), )
    region1=regionToolset.Region(side1Faces=side1Faces1)
    a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['Steel-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#8 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].SurfaceToSurfaceContactExp(name ='Int-3', 
        createStepName='Initial', master = region1, slave = region2, 
        mechanicalConstraint=KINEMATIC, sliding=FINITE, 
        interactionProperty='IntProp-1', initialClearance=OMIT, datumAxis=None, 
        clearanceRegion=None)
    #: The interaction "Int-3" has been created.
    mdb.models[ModleName].Interaction(name='Int-3-Copy', 
        objectToCopy=mdb.models[ModleName].interactions['Int-3'], 
        toStepName='Initial')
    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['Concrete-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#4 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].interactions['Int-3-Copy'].setValues(slave=region2, 
        mechanicalConstraint=KINEMATIC, sliding=FINITE, 
        interactionProperty='IntProp-1', initialClearance=OMIT, datumAxis=None, 
        clearanceRegion=None)

    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['Concrete-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#1 ]', ), )
    region1=regionToolset.Region(side1Faces=side1Faces1)
    #a = mdb.models[ModleName].rootAssembly
    s1 = a.instances['Steel-1'].faces
    side1Faces1 = s1.getSequenceFromMask(mask=('[#2 ]', ), )
    region2=regionToolset.Region(side1Faces=side1Faces1)
    mdb.models[ModleName].SurfaceToSurfaceContactExp(name ='Int-5', 
        createStepName='Initial', master = region1, slave = region2, 
        mechanicalConstraint=KINEMATIC, sliding=FINITE, 
        interactionProperty='IntProp-1', initialClearance=OMIT, datumAxis=None, 
        clearanceRegion=None)
    #: The interaction "Int-5" has been created.



#coupling
    #a = mdb.models[ModleName].rootAssembly
    f1 = a.instances['plate-1'].faces
    faces1 = f1.getSequenceFromMask(mask=('[#3f ]', ), )
    region2=regionToolset.Region(faces=faces1)
    #a = mdb.models[ModleName].rootAssembly
    r1 = a.referencePoints
    refPoints1=(r1[11], )
    region1=regionToolset.Region(referencePoints=refPoints1)
    mdb.models[ModleName].RigidBody(name='Constraint-1', refPointRegion=region1, 
        bodyRegion=region2)
    #a = mdb.models[ModleName].rootAssembly
    f1 = a.instances['plate-1-lin-2-1'].faces
    faces1 = f1.getSequenceFromMask(mask=('[#3f ]', ), )
    region2=regionToolset.Region(faces=faces1)
    #a = mdb.models[ModleName].rootAssembly
    r1 = a.referencePoints
    refPoints1=(r1[10], )
    region1=regionToolset.Region(referencePoints=refPoints1)
    mdb.models[ModleName].RigidBody(name='Constraint-2', refPointRegion=region1, 
        bodyRegion=region2)

#load
    mdb.models[ModleName].TabularAmplitude(name='Amp-1', timeSpan=STEP, 
    smooth=SOLVER_DEFAULT, data=((0.0, 0.0), (1.0, 1.0)))
    #a = mdb.models[ModleName].rootAssembly
    r1 = a.referencePoints
    refPoints1=(r1[11], )
    region = regionToolset.Region(referencePoints=refPoints1)
    mdb.models[ModleName].DisplacementBC(name='BC-1', createStepName='Initial', 
        region=region, u1=SET, u2=SET, u3=SET, ur1=SET, ur2=SET, ur3=SET, 
        amplitude=UNSET, distributionType=UNIFORM, fieldName='', localCsys=None)

    #a = mdb.models[ModleName].rootAssembly
    r1 = a.referencePoints
    refPoints1=(r1[10], )
    region = regionToolset.Region(referencePoints=refPoints1)
    mdb.models[ModleName].DisplacementBC(name='BC-2', createStepName='Step-1', 
        region=region, u1=0.0, u2=0.0, u3=-15.0, ur1=0.0, ur2=0.0, ur3=0.0, 
        amplitude='Amp-1', fixed=OFF, distributionType=UNIFORM, fieldName='', 
        localCsys=None)

    ##Mesh
    #partition cell
    p = mdb.models[ModleName].parts['Steel']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    e, v1, d1 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e[0], point=v1[0], cells=pickedCells)
    # p = mdb.models[ModleName].parts['Steel']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#3 ]', ), )
    e1, v2, d2 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e1[9], cells=pickedCells, 
        point=p.InterestingPoint(edge=e1[9], rule=MIDDLE))

    p = mdb.models[ModleName].parts['Concrete']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    e, v1, d1 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e[0], cells=pickedCells, 
        point=p.InterestingPoint(edge=e[0], rule=MIDDLE))
    # p = mdb.models[ModleName].parts['Concrete']
    c = p.cells
    pickedCells = c.getSequenceFromMask(mask=('[#3 ]', ), )
    e1, v2, d2 = p.edges, p.vertices, p.datums
    p.PartitionCellByPlaneNormalToEdge(edge=e1[4], cells=pickedCells, 
        point=p.InterestingPoint(edge=e1[4], rule=MIDDLE))
    # draw mesh
    p = mdb.models[ModleName].parts['Concrete']
    p.seedPart(size=10.0, deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models[ModleName].parts['Concrete']
    p.generateMesh()
    p = mdb.models[ModleName].parts['Steel']
    p.seedPart(size=10.0, deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models[ModleName].parts['Steel']
    e = p.edges
    pickedEdges = e.getSequenceFromMask(mask=('[#c000041 ]', ), )
    p.seedEdgeBySize(edges=pickedEdges, size=5.0, deviationFactor=0.1, 
    minSizeFactor=0.1, constraint=FINER)
    p.generateMesh()
    p = mdb.models[ModleName].parts['plate']
    p.seedPart(size=200.0, deviationFactor=0.1, minSizeFactor=0.1)
    p = mdb.models[ModleName].parts['plate']
    p.generateMesh()

wb.save(filename)