from errno import ECONNABORTED
from telnetlib import EC
from abaqus import *
from abaqusConstants import *
from odbAccess import *
from textRepr  import *
import regionToolset
import openpyxl as op
filename=r'E:\DOCUMENTS\1YueLu\1.Subjects\Poisson\Data\DataBase\Uniaxial_Biaxial_Triaxial_Test.xlsx'
wb=op.load_workbook(filename)
sh=wb["ABAQUSDataBase"]


for i in range(55,62): #excel row-1  : end row **********input*******
    #-----Create a new model
    ModelName=str(sh.cell(row=i,column=1).value)
    Dimensions=round(sh.cell(row=i,column=5).value,2)
    Radius=Dimensions/2.
    Highth=round(sh.cell(row=i,column=6).value,2)  #TubeLength

    fl=round(sh.cell(row=i,column=2).value,2) #letral presures
    Fc=round(sh.cell(row=i,column=3).value,2)  #Concrete strength
    Econ=round(sh.cell(row=i,column=4).value,2)
    
    ft=Fc/10.
    

    mdb.Model(name=ModelName, modelType=STANDARD_EXPLICIT)
    ###Part
    s = mdb.models[ModelName].ConstrainedSketch(name='__profile__', 
        sheetSize=200.0)
    g, v, d, c = s.geometry, s.vertices, s.dimensions, s.constraints
    s.setPrimaryObject(option=STANDALONE)
    s.rectangle(point1=(50.0, 50.0), point2=(-50.0, -50.0))
    p = mdb.models[ModelName].Part(name='concrete', dimensionality=THREE_D, 
        type=DEFORMABLE_BODY)
    p = mdb.models[ModelName].parts['concrete']
    p.BaseSolidExtrude(sketch=s, depth=100)
    s.unsetPrimaryObject()
    p = mdb.models[ModelName].parts['concrete']


    ##property
    mdb.models[ModelName].Material(name='VUmat')
    mdb.models[ModelName].materials['VUmat'].Density(table=((2.4e-09, ), ))
    mdb.models[ModelName].materials['VUmat'].Depvar(deleteVar=28, n=49)
    if fl>0:
        mdb.models[ModelName].materials['VUmat'].UserMaterial(mechanicalConstants=(1.0, 
            30000, 0.2, Fc, ft, 0.01, 0.0, 0.08, 0.003, 2.0, 1e-06, 0.01, 15.0))
    else:
         mdb.models[ModelName].materials['VUmat'].UserMaterial(mechanicalConstants=(1.0, 
            30000, 0.2, Fc, ft, 0.01, 0.0, 0.08, 0.003, 2.0, 1e-06, 0.01, 5.0))
    mdb.models[ModelName].HomogeneousSolidSection(name='Concrete', 
        material='VUmat', thickness=None)
    p = mdb.models[ModelName].parts['concrete']
    c = p.cells
    cells = c.getSequenceFromMask(mask=('[#1 ]', ), )
    region = regionToolset.Region(cells=cells)
    p = mdb.models[ModelName].parts['concrete']
    p.SectionAssignment(region=region, sectionName='Concrete', offset=0.0, 
        offsetType=MIDDLE_SURFACE, offsetField='', 
        thicknessAssignment=FROM_SECTION)



    ##--------

    ##Assemble
    a = mdb.models[ModelName].rootAssembly
    a.DatumCsysByDefault(CARTESIAN)
    p = mdb.models[ModelName].parts['concrete']
    a.Instance(name='concrete-1', part=p, dependent=ON)

    # a.makeIndependent(instances=(a.instances['concrete-1'], ))

    # c1 = a.instances['concrete-1'].cells
    # pickedCells = c1.getSequenceFromMask(mask=('[#1 ]', ), )
    # v11 = a.instances['concrete-1'].vertices
    # d11 = a.datums
    # a.PartitionCellByPlanePointNormal(point=v11[0], normal=d11[1].axis2, 
    #     cells=pickedCells)

    # c1 = a.instances['concrete-1'].cells
    # pickedCells = c1.getSequenceFromMask(mask=('[#3 ]', ), )
    # e1 = a.instances['concrete-1'].edges
    # d21 = a.datums
    # a.PartitionCellByPlanePointNormal(normal=d21[1].axis1, cells=pickedCells, 
    #     point=a.instances['concrete-1'].InterestingPoint(edge=e1[4], rule=MIDDLE))

     
    if fl>0:
        ##Step
        mdb.models[ModelName].ExplicitDynamicsStep(name='Step-1', previous='Initial', 
            massScaling=((SEMI_AUTOMATIC, MODEL, AT_BEGINNING, 0.0, 0.0001, BELOW_MIN, 
            0, 0, 0.0, 0.0, 0, None), ), improvedDtMethod=ON)
        
        mdb.models[ModelName].fieldOutputRequests['F-Output-1'].setValues(variables=(
            'S', 'SVAVG', 'PE', 'PEVAVG', 'PEEQ', 'PEEQVAVG', 'LE', 'U', 'V', 'A', 
            'RF', 'CSTRESS', 'EVF', 'SDV', 'STATUS','EVOL'), numIntervals=100)

        mdb.models[ModelName].ExplicitDynamicsStep(name='Step-2', previous='Step-1', 
            massScaling=((SEMI_AUTOMATIC, MODEL, AT_BEGINNING, 0.0, 0.0001, BELOW_MIN, 
            0, 0, 0.0, 0.0, 0, None), ), improvedDtMethod=ON)
        


        #Assemnle
        a = mdb.models[ModelName].rootAssembly
        a.ReferencePoint(point=(0.0, 0.0, -20.0))

        a.ReferencePoint(point=(0.0, 0.0, Highth+20))


        r1 = a.referencePoints
        refPoints1=(r1[5], )
        region1=regionToolset.Region(referencePoints=refPoints1)

        s1 = a.instances['concrete-1'].faces
        side1Faces1 = s1.getSequenceFromMask(mask=('[#10 ]', ), )
        region2=regionToolset.Region(side1Faces=side1Faces1)
        mdb.models[ModelName].Coupling(name='Constraint-1', controlPoint=region1, 
            surface=region2, influenceRadius=WHOLE_SURFACE, couplingType=KINEMATIC, 
            localCsys=None, u1=OFF, u2=OFF, u3=ON, ur1=ON, ur2=ON, ur3=ON)


        


        #Load

        mdb.models[ModelName].TabularAmplitude(name='Amp-1', timeSpan=STEP, 
            smooth=SOLVER_DEFAULT, data=((0.0, 0.0), (1.0, 1.0)))
        mdb.models[ModelName].SmoothStepAmplitude(name='Amp-2', timeSpan=STEP, data=((
            0.0, 0.0), (1.0, 1.0)))
        a = mdb.models[ModelName].rootAssembly
        s1 = a.instances['concrete-1'].faces
        side1Faces1 = s1.getSequenceFromMask(mask=('[#f ]', ), )
        region = regionToolset.Region(side1Faces=side1Faces1)
        mdb.models[ModelName].Pressure(name='Load-1', createStepName='Step-1', 
            region=region, distributionType=UNIFORM, field='', magnitude=fl, 
            amplitude='Amp-2')

        a = mdb.models[ModelName].rootAssembly
        f1 = a.instances['concrete-1'].faces
        faces1 = f1.getSequenceFromMask(mask=('[#20 ]', ), )
        region = regionToolset.Region(faces=faces1)
        mdb.models[ModelName].DisplacementBC(name='BC-1', createStepName='Initial', 
            region=region, u1=UNSET, u2=UNSET, u3=SET, ur1=SET, ur2=SET, ur3=SET, 
            amplitude=UNSET, distributionType=UNIFORM, fieldName='', localCsys=None)
        # mdb.models[ModelName].boundaryConditions['BC-1'].setValuesInStep(
        #     stepName='Step-2', u1=FREED, u2=FREED)
        

        a = mdb.models[ModelName].rootAssembly
        r1 = a.referencePoints
        refPoints1=(r1[5], )
        region = regionToolset.Region(referencePoints=refPoints1)
        mdb.models[ModelName].DisplacementBC(name='BC-2', createStepName='Step-1', 
            region=region, u1=UNSET, u2=UNSET, u3=0.0, ur1=UNSET, ur2=UNSET, ur3=UNSET, 
            amplitude=UNSET, fixed=OFF,distributionType=UNIFORM, fieldName='', localCsys=None)
        mdb.models[ModelName].boundaryConditions['BC-2'].deactivate('Step-2')
        
        
        mdb.models[ModelName].DisplacementBC(name='BC-3', createStepName='Step-2', 
            region=region, u1=0.0, u2=0.0, u3=-6.0, ur1=0.0, ur2=0.0, ur3=0.0, 
            amplitude='Amp-1', fixed=OFF, distributionType=UNIFORM, fieldName='', 
            localCsys=None)

      
    else:
        ##Step
        mdb.models[ModelName].ExplicitDynamicsStep(name='Step-1', previous='Initial', 
            massScaling=((SEMI_AUTOMATIC, MODEL, AT_BEGINNING, 0.0, 0.0001, BELOW_MIN, 
            0, 0, 0.0, 0.0, 0, None), ), improvedDtMethod=ON)
        
        mdb.models[ModelName].fieldOutputRequests['F-Output-1'].setValues(variables=(
            'S', 'SVAVG', 'PE', 'PEVAVG', 'PEEQ', 'PEEQVAVG', 'LE', 'U', 'V', 'A', 
            'RF', 'CSTRESS', 'EVF', 'SDV', 'STATUS','EVOL'), numIntervals=100)




        #Assemnle
        a = mdb.models[ModelName].rootAssembly
        a.ReferencePoint(point=(0.0, 0.0, -20.0))

        a.ReferencePoint(point=(0.0, 0.0, 100+20))


        r1 = a.referencePoints
        refPoints1=(r1[5], )
        region1=regionToolset.Region(referencePoints=refPoints1)

        s1 = a.instances['concrete-1'].faces
        side1Faces1 = s1.getSequenceFromMask(mask=('[#10 ]', ), )
        region2=regionToolset.Region(side1Faces=side1Faces1)
        mdb.models[ModelName].Coupling(name='Constraint-1', controlPoint=region1, 
            surface=region2, influenceRadius=WHOLE_SURFACE, couplingType=KINEMATIC, 
            localCsys=None, u1=OFF, u2=OFF, u3=ON, ur1=ON, ur2=ON, ur3=ON)


    


        #Load

        mdb.models[ModelName].TabularAmplitude(name='Amp-1', timeSpan=STEP, 
            smooth=SOLVER_DEFAULT, data=((0.0, 0.0), (1.0, 1.0)))
        
        a = mdb.models[ModelName].rootAssembly

        a = mdb.models[ModelName].rootAssembly
        f1 = a.instances['concrete-1'].faces
        faces1 = f1.getSequenceFromMask(mask=('[#20 ]', ), )
        region = regionToolset.Region(faces=faces1)
        mdb.models[ModelName].DisplacementBC(name='BC-1', createStepName='Initial', 
            region=region, u1=UNSET, u2=UNSET, u3=SET, ur1=SET, ur2=SET, ur3=SET, 
            amplitude=UNSET, distributionType=UNIFORM, fieldName='', localCsys=None)

        a = mdb.models[ModelName].rootAssembly
        r1 = a.referencePoints
        refPoints1=(r1[5], )
        region = regionToolset.Region(referencePoints=refPoints1)
        mdb.models[ModelName].DisplacementBC(name='BC-2', createStepName='Step-1', 
            region=region, u1=0.0, u2=0.0, u3=-6, ur1=0.0, ur2=0.0, ur3=0.0, 
            amplitude='Amp-1', fixed=OFF, distributionType=UNIFORM, fieldName='', 
            localCsys=None)






    #Mesh
    a = mdb.models[ModelName].rootAssembly
    p = mdb.models[ModelName].parts['concrete']
    p.seedPart(size=6.0, deviationFactor=0.1, minSizeFactor=0.1)
    p.generateMesh()
   

    #regenerate
    a = mdb.models[ModelName].rootAssembly
    a.regenerate()


